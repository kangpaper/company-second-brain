import csv
import html
import re
import zipfile
from dataclasses import dataclass
from html.parser import HTMLParser
from io import BytesIO, StringIO
from typing import Any

import pymupdf
from docx import Document as open_docx
from openpyxl import load_workbook

MAX_INPUT_BYTES = 10 * 1024 * 1024
MAX_EXPANDED_BYTES = 50 * 1024 * 1024
MAX_TEXT_CHARS = 2_000_000
MAX_CANDIDATES = 5_000
MAX_PDF_PAGES = 200


class ParseError(ValueError):
    pass


@dataclass(frozen=True)
class ExtractionCandidateData:
    kind: str
    locator: dict[str, Any]
    data: dict[str, Any]
    text: str


@dataclass(frozen=True)
class ParsedContent:
    text: str
    candidates: list[ExtractionCandidateData]
    metadata: dict[str, Any]


def ensure_candidate_capacity(candidates: list[ExtractionCandidateData]) -> None:
    if len(candidates) >= MAX_CANDIDATES:
        raise ParseError("Extraction candidate count exceeds 5000")


def decoded_text(content: bytes) -> str:
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise ParseError("Text input must be UTF-8") from error


def ensure_archive_safe(content: bytes) -> None:
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            expanded = sum(info.file_size for info in archive.infolist())
            if expanded > MAX_EXPANDED_BYTES:
                raise ParseError("Office document expanded size exceeds 50 MiB")
    except zipfile.BadZipFile as error:
        raise ParseError("Invalid Office document") from error


def bounded_result(
    text_parts: list[str],
    candidates: list[ExtractionCandidateData],
    metadata: dict[str, Any],
) -> ParsedContent:
    text = "\n".join(part.strip() for part in text_parts if part.strip()).strip()
    if not text:
        raise ParseError("Document contains no extractable text")
    if len(text) > MAX_TEXT_CHARS:
        raise ParseError("Extracted text exceeds 2000000 characters")
    if len(candidates) > MAX_CANDIDATES:
        raise ParseError("Extraction candidate count exceeds 5000")
    return ParsedContent(text=text, candidates=candidates, metadata=metadata)


def parse_markdown(
    content: bytes,
) -> tuple[list[str], list[ExtractionCandidateData], dict[str, Any]]:
    source = decoded_text(content)
    parts: list[str] = []
    candidates: list[ExtractionCandidateData] = []
    heading = "Document"
    section_lines: list[str] = []
    section_start = 1

    def emit() -> None:
        nonlocal section_lines
        text = "\n".join(section_lines).strip()
        if text:
            ensure_candidate_capacity(candidates)
            parts.append(text)
            candidates.append(
                ExtractionCandidateData(
                    kind="section",
                    locator={"heading": heading, "start_line": section_start},
                    data={"heading": heading},
                    text=text,
                )
            )
        section_lines = []

    for line_number, line in enumerate(source.splitlines(), start=1):
        match = re.match(r"^#{1,6}\s+(.+)$", line)
        if match:
            emit()
            heading = match.group(1).strip()
            section_start = line_number
        else:
            section_lines.append(line)
    emit()
    return parts, candidates, {"format": "markdown"}


def parse_plain_text(
    content: bytes,
) -> tuple[list[str], list[ExtractionCandidateData], dict[str, Any]]:
    source = decoded_text(content)
    parts: list[str] = []
    candidates: list[ExtractionCandidateData] = []
    section_lines: list[str] = []
    section_start = 1

    def emit() -> None:
        nonlocal section_lines
        text = "\n".join(section_lines).strip()
        if text:
            ensure_candidate_capacity(candidates)
            parts.append(text)
            candidates.append(
                ExtractionCandidateData(
                    kind="section",
                    locator={"start_line": section_start},
                    data={},
                    text=text,
                )
            )
        section_lines = []

    for line_number, line in enumerate(source.splitlines(), start=1):
        if not line.strip():
            emit()
            section_start = line_number + 1
        else:
            section_lines.append(line)
    emit()
    return parts, candidates, {"format": "text"}


def parse_csv(content: bytes) -> tuple[list[str], list[ExtractionCandidateData], dict[str, Any]]:
    reader = csv.DictReader(StringIO(decoded_text(content), newline=""))
    if not reader.fieldnames or any(
        field is None or not field.strip() for field in reader.fieldnames
    ):
        raise ParseError("CSV requires non-empty headers")
    parts: list[str] = []
    candidates: list[ExtractionCandidateData] = []
    for row_number, row in enumerate(reader, start=2):
        ensure_candidate_capacity(candidates)
        data = {str(key): value for key, value in row.items()}
        text = " | ".join(f"{key}: {value}" for key, value in data.items())
        parts.append(text)
        candidates.append(
            ExtractionCandidateData(
                kind="row", locator={"row": row_number}, data=data, text=text
            )
        )
    return parts, candidates, {"format": "csv", "columns": reader.fieldnames}


def cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def unique_headers(headers: list[str]) -> list[str]:
    result: list[str] = []
    used: set[str] = set()
    for index, header in enumerate(headers, start=1):
        base = header or f"column_{index}"
        candidate = base
        suffix = 2
        while candidate in used:
            candidate = f"{base}_{suffix}"
            suffix += 1
        used.add(candidate)
        result.append(candidate)
    return result


def parse_xlsx(content: bytes) -> tuple[list[str], list[ExtractionCandidateData], dict[str, Any]]:
    ensure_archive_safe(content)
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise ParseError("Invalid XLSX document") from error
    parts: list[str] = []
    candidates: list[ExtractionCandidateData] = []
    sheet_names: list[str] = []
    try:
        for sheet in workbook.worksheets:
            sheet_names.append(sheet.title)
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if headers is None:
                continue
            names = [str(value).strip() if value is not None else "" for value in headers]
            if any(not name for name in names) or len(names) != len(set(names)):
                raise ParseError(f"Worksheet {sheet.title} requires unique non-empty headers")
            for row_number, values in enumerate(rows, start=2):
                data = {name: cell_value(value) for name, value in zip(names, values, strict=True)}
                if not any(value not in (None, "") for value in data.values()):
                    continue
                ensure_candidate_capacity(candidates)
                text = " | ".join(f"{key}: {value}" for key, value in data.items())
                parts.append(text)
                candidates.append(
                    ExtractionCandidateData(
                        kind="row",
                        locator={"sheet": sheet.title, "row": row_number},
                        data=data,
                        text=text,
                    )
                )
    finally:
        workbook.close()
    return parts, candidates, {"format": "xlsx", "sheets": sheet_names}


def parse_docx(content: bytes) -> tuple[list[str], list[ExtractionCandidateData], dict[str, Any]]:
    ensure_archive_safe(content)
    try:
        document = open_docx(BytesIO(content))
    except Exception as error:
        raise ParseError("Invalid DOCX document") from error
    parts: list[str] = []
    candidates: list[ExtractionCandidateData] = []
    for paragraph_number, paragraph in enumerate(document.paragraphs, start=1):
        text = paragraph.text.strip()
        if text:
            ensure_candidate_capacity(candidates)
            parts.append(text)
            candidates.append(
                ExtractionCandidateData(
                    kind="paragraph",
                    locator={"paragraph": paragraph_number},
                    data={"style": paragraph.style.name if paragraph.style else None},
                    text=text,
                )
            )
    for table_number, table in enumerate(document.tables, start=1):
        raw_headers = [cell.text.strip() for cell in table.rows[0].cells] if table.rows else []
        headers = unique_headers(raw_headers)
        for row_number, row in enumerate(table.rows[1:], start=2):
            values = [cell.text.strip() for cell in row.cells]
            if not any(values):
                continue
            data = dict(zip(headers, values, strict=False))
            text = " | ".join(f"{key}: {value}" for key, value in data.items() if value)
            if text:
                ensure_candidate_capacity(candidates)
                parts.append(text)
                candidates.append(
                    ExtractionCandidateData(
                        kind="table_row",
                        locator={"table": table_number, "row": row_number},
                        data=data,
                        text=text,
                    )
                )
    return parts, candidates, {"format": "docx"}


def parse_pdf(content: bytes) -> tuple[list[str], list[ExtractionCandidateData], dict[str, Any]]:
    try:
        document: Any = pymupdf.open(  # type: ignore[no-untyped-call]
            stream=content, filetype="pdf"
        )
    except Exception as error:
        raise ParseError("Invalid PDF document") from error
    parts: list[str] = []
    candidates: list[ExtractionCandidateData] = []
    try:
        if document.page_count > MAX_PDF_PAGES:
            raise ParseError("PDF exceeds 200 pages")
        for index, page in enumerate(document, start=1):
            text = page.get_text("text").strip()
            if text:
                parts.append(text)
                candidates.append(
                    ExtractionCandidateData(
                        kind="page", locator={"page": index}, data={}, text=text
                    )
                )
    finally:
        document.close()
    return parts, candidates, {"format": "pdf", "pages": len(candidates)}


class VisibleHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.hidden_depth = 0
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"script", "style", "noscript"}:
            self.hidden_depth += 1
        if tag in {"p", "div", "article", "section", "h1", "h2", "h3", "li", "br"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style", "noscript"} and self.hidden_depth:
            self.hidden_depth -= 1
        if tag in {"p", "div", "article", "section", "h1", "h2", "h3", "li"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if not self.hidden_depth:
            self.parts.append(data)


def parse_html(content: bytes) -> tuple[list[str], list[ExtractionCandidateData], dict[str, Any]]:
    parser = VisibleHTMLParser()
    try:
        parser.feed(decoded_text(content))
    except Exception as error:
        raise ParseError("Invalid HTML document") from error
    normalized = html.unescape("".join(parser.parts))
    sections = [" ".join(part.split()) for part in re.split(r"\n+", normalized) if part.strip()]
    if len(sections) > MAX_CANDIDATES:
        raise ParseError("Extraction candidate count exceeds 5000")
    candidates = [
        ExtractionCandidateData(
            kind="section", locator={"section": index}, data={}, text=text
        )
        for index, text in enumerate(sections, start=1)
    ]
    return sections, candidates, {"format": "html"}


PARSERS = {
    "text/markdown": parse_markdown,
    "text/plain": parse_plain_text,
    "text/csv": parse_csv,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": parse_xlsx,
    "application/pdf": parse_pdf,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": parse_docx,
    "text/html": parse_html,
}


def parse_content(media_type: str, filename: str, content: bytes) -> ParsedContent:
    if len(content) > MAX_INPUT_BYTES:
        raise ParseError("Input exceeds 10 MiB")
    parser = PARSERS.get(media_type.lower().split(";", maxsplit=1)[0].strip())
    if parser is None:
        raise ParseError(f"Unsupported media type: {media_type}")
    parts, candidates, metadata = parser(content)
    metadata = {**metadata, "filename": filename, "byte_size": len(content)}
    return bounded_result(parts, candidates, metadata)
